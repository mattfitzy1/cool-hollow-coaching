"""
Builds one filled-in demo sample per milestone tool, all describing the same
fictional client so a team walkthrough tells one coherent story end to end:

Palmetto Air & Plumbing, a $4.2M/year HVAC and plumbing company outside
Charleston SC. Owner-operator Dave Whitfield is the classic single point of
failure: approves every quote, is the only one who can price complex jobs,
and the business can't run a day without him.

Each sample is built on top of the CURRENT branded template (Instructions,
Examples, Starter Ideas, data tab(s)) with Dave's data written into the
data tab(s) only, so uploading it into its tool is the exact experience a
real client will have. Each file gets its own milestone-specific name
(not a generic "Sample - Filled In.xlsx" repeated seven times), so several
can be open in Excel at once without name collisions.

Usage:
    apps/hidden-profit-analyzer/.venv/bin/python3 apps/_shared/generate_samples.py
"""

from pathlib import Path
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
CLIENT = "Palmetto Air and Plumbing"


def sample_name(milestone_number: int, tool_name: str) -> str:
    return f"Sample - Milestone {milestone_number} {tool_name} - {CLIENT}.xlsx"


def _fill_tab(wb, tab_name, rows, start_row=5):
    ws = wb[tab_name]
    for r_off, row in enumerate(rows):
        for c_off, val in enumerate(row):
            ws.cell(row=start_row + r_off, column=1 + c_off, value=val)


def build_reclaim_protocol():
    template = REPO_ROOT / "apps/reclaim-protocol/Cool_Hollow_Coaching_Milestone_1_Reclaim_Protocol_Template.xlsx"
    wb = load_workbook(template)
    _fill_tab(wb, "Time Log", [
        ["Approve every HVAC and plumbing quote before it goes out", 6, "delegate", "yes", "yes", "low"],
        ["Chase overdue customer invoices", 3, "automate", "no", "no", "medium"],
        ["Run payroll every two weeks", 2, "delegate", "yes", "no", "medium"],
        ["Sit in on every new technician interview", 3, "delegate", "no", "no", "low"],
        ["Price every complex commercial job", 5, "owner_only", "no", "no", "high"],
        ["Answer after-hours emergency service calls", 4, "eliminate", "no", "no", "high"],
        ["Post to Instagram and Facebook for the business", 2, "automate", "no", "no", "low"],
        ["Reconcile bank statements", 2, "delegate", "yes", "yes", "low"],
        ["Final review of every commercial proposal", 4, "delegate", "yes", "no", "medium"],
        ["Set next year's pricing strategy", 3, "owner_only", "no", "no", "high"],
    ])
    out = REPO_ROOT / "apps/reclaim-protocol" / sample_name(1, "Reclaim Protocol")
    wb.save(out)
    return out


def build_impact_map():
    template = REPO_ROOT / "apps/impact-map/Cool_Hollow_Coaching_Milestone_2_Impact_Map_Template.xlsx"
    wb = load_workbook(template)
    _fill_tab(wb, "Initiative List", [
        ["Launch a maintenance membership program", 5, 4, 5, 3],
        ["Open a second truck bay and expand the fleet", 3, 2, 4, 4],
        ["Redesign the online booking system", 4, 4, 4, 3],
        ["Sponsor the county fair", 1, 1, 1, 2],
        ["Build a referral network with local realtors", 4, 5, 4, 2],
        ["Add a new commercial HVAC division", 3, 3, 5, 5],
        ["Rebrand the trucks and uniforms", 2, 2, 2, 3],
        ["Implement a new CRM", 4, 3, 4, 4],
    ])
    out = REPO_ROOT / "apps/impact-map" / sample_name(2, "Impact Map")
    wb.save(out)
    return out


def build_dashboard_selector():
    template = REPO_ROOT / "apps/dashboard-selector/Cool_Hollow_Coaching_Milestone_3_Dashboard_Template.xlsx"
    wb = load_workbook(template)
    _fill_tab(wb, "Candidate Metrics", [
        ["cash", "13-week cash runway (weeks)", 7, 13, "higher_better", "yes", 5],
        ["cash", "Days sales outstanding", 42, 25, "lower_better", "yes", 4],
        ["sales", "Monthly revenue", 350000, 400000, "higher_better", "no", 5],
        ["sales", "Close rate on quotes", 38, 55, "higher_better", "yes", 4],
        ["delivery", "On-time completion rate", 82, 95, "higher_better", "yes", 5],
        ["delivery", "Callback / rework rate", 9, 3, "lower_better", "no", 3],
        ["customer", "Net promoter score", 44, 65, "higher_better", "yes", 4],
        ["customer", "Google review rating", 4.1, 4.7, "higher_better", "no", 3],
        ["team", "Technician turnover, last 90 days", 3, 0, "lower_better", "no", 4],
        ["team", "Open technician roles", 2, 0, "lower_better", "no", 3],
    ])
    out = REPO_ROOT / "apps/dashboard-selector" / sample_name(3, "Dashboard")
    wb.save(out)
    return out


def build_profit_discovery_audit():
    template = REPO_ROOT / "apps/profit-discovery-audit/Cool_Hollow_Coaching_Milestone_4_Profit_Discovery_Audit_Template.xlsx"
    wb = load_workbook(template)
    _fill_tab(wb, "P&L", [
        ["Revenue", 340000, 355000, 368000],
        ["Cost of Goods Sold", 187000, 195000, 202000],
        ["Payroll", 78000, 79000, 80000],
        ["Rent", 6500, 6500, 6500],
        ["Vehicle Fuel and Maintenance", 9200, 9800, 9500],
        ["Software Subscriptions", 1100, 1150, 3200],
        ["Marketing", 5200, 5400, 5600],
        ["Discounts Given", 4800, 5100, 5300],
        ["Insurance", 3100, 3100, 3100],
    ])
    _fill_tab(wb, "Customer Service Breakdown", [
        ["customer", "Meridian Property Group", 48000, 41000],
        ["customer", "Coastal Retail Plaza", 32000, 19000],
        ["customer", "Sunridge HOA", 27000, 15000],
        ["service", "Emergency Service Calls", 62000, 28000],
        ["service", "Maintenance Membership Plan", 54000, 18000],
        ["service", "New Commercial Installs", 95000, 71000],
    ])
    out = REPO_ROOT / "apps/profit-discovery-audit" / sample_name(4, "Profit Discovery Audit")
    wb.save(out)
    return out


def build_cash_confidence():
    template = REPO_ROOT / "apps/cash-confidence/Cool_Hollow_Coaching_Milestone_5_Cash_Confidence_Template.xlsx"
    wb = load_workbook(template)
    _fill_tab(wb, "Cash Items", [
        [1, "inflow", "client revenue", "Recurring service contracts", 78000],
        [1, "outflow", "Payroll", "Biweekly payroll", 39000],
        [1, "outflow", "Vehicle Fuel and Maintenance", "Fleet fuel and repairs", 4200],
        [2, "inflow", "client revenue", "Commercial install draw", 45000],
        [2, "outflow", "rent", "Shop lease", 6500],
        [3, "outflow", "Payroll", "Biweekly payroll", 39000],
        [3, "outflow", "Software Subscriptions", "Annual CRM renewal", 3200],
        [4, "inflow", "client revenue", "Recurring service contracts", 81000],
        [5, "outflow", "Payroll", "Biweekly payroll", 39000],
        [6, "inflow", "client revenue", "Commercial install draw", 52000],
        [7, "outflow", "Payroll", "Biweekly payroll", 39000],
        [7, "outflow", "insurance", "Annual liability insurance renewal", 11000],
        [8, "inflow", "client revenue", "Recurring service contracts", 79000],
        [9, "outflow", "Payroll", "Biweekly payroll", 39000],
        [10, "inflow", "client revenue", "Commercial install draw", 48000],
        [11, "outflow", "Payroll", "Biweekly payroll", 39000],
        [12, "outflow", "quarterly tax payment", "Estimated quarterly taxes", 22000],
        [13, "inflow", "client revenue", "Recurring service contracts", 83000],
        [13, "outflow", "Payroll", "Biweekly payroll", 39000],
    ])
    _fill_tab(wb, "Recurring Expenses", [
        ["Payroll", 19500, 5, 5, 5, 3, 5],
        ["Software Subscriptions", 250, 3, 2, 3, 2, 3],
        ["Fleet Fuel and Maintenance", 2100, 4, 3, 4, 3, 4],
        ["Underused lead-gen subscription", 180, 1, 1, 1, 1, 1],
        ["Shop Lease", 1500, 4, 4, 5, 2, 5],
    ])
    _fill_tab(wb, "Receivables Aging", [
        ["Meridian Property Group", 22000, 30, 68],
        ["Coastal Retail Plaza", 9500, 30, 22],
        ["Sunridge HOA", 14000, 15, 51],
    ])
    out = REPO_ROOT / "apps/cash-confidence" / sample_name(5, "Cash Confidence")
    wb.save(out)
    return out


def build_bottleneck_breakthrough():
    template = REPO_ROOT / "apps/bottleneck-breakthrough/Cool_Hollow_Coaching_Milestone_6_Bottleneck_Breakthrough_Template.xlsx"
    wb = load_workbook(template)
    _fill_tab(wb, "Constraint Worksheet", [
        ["Dave approves every quote before it goes out", "Sales", 18, 0.4, 5, "yes", 2],
        ["Only Dave can price complex commercial jobs", "Sales", 4, 1.5, 5, "no", 5],
        ["Techs wait on dispatch for their next job", "Delivery", 25, 0.3, 4, "yes", 2],
        ["Invoices built by hand from job tickets", "Billing", 15, 0.5, 4, "yes", 1],
        ["No documented onboarding for new technicians", "Team", 1, 6, 3, "no", 4],
    ])
    out = REPO_ROOT / "apps/bottleneck-breakthrough" / sample_name(6, "Bottleneck Breakthrough")
    wb.save(out)
    return out


def build_team_builder():
    template = REPO_ROOT / "apps/team-builder/Cool_Hollow_Coaching_Milestone_7_Team_Builder_Template.xlsx"
    wb = load_workbook(template)
    _fill_tab(wb, "Roles", [
        ["Operations Manager", 2, 5, "Keeps every job on schedule",
         "Owns vendor and supplier relationships", "Runs the weekly dispatch huddle",
         "Can approve spend up to $3,000 without sign-off", "On-time completion at or above 95%"],
        ["Sales Lead", 2, 4, "Approves standard quotes without Dave",
         "Follows up every open proposal within 48 hours", "Owns the close-rate number",
         "Can discount up to 10% without approval", "Close rate at or above 55%"],
        ["Office Manager", 3, 5, "Owns invoicing and collections",
         "Runs payroll every cycle", "Keeps the books reconciled monthly",
         "Can chase and settle invoices without Dave", "Days sales outstanding at or below 25"],
    ])
    _fill_tab(wb, "Candidates", [
        ["Maria Alvarez", "Operations Manager", 5, 4, 5, 4, 4],
        ["Kevin Boyd", "Operations Manager", 3, 5, 4, 5, 3],
        ["Renee Foster", "Sales Lead", 4, 5, 4, 4, 5],
    ])
    out = REPO_ROOT / "apps/team-builder" / sample_name(7, "Team Builder")
    wb.save(out)
    return out


if __name__ == "__main__":
    builders = [
        build_reclaim_protocol, build_impact_map, build_dashboard_selector,
        build_profit_discovery_audit, build_cash_confidence,
        build_bottleneck_breakthrough, build_team_builder,
    ]
    for builder in builders:
        path = builder()
        print(f"Built: {path.relative_to(REPO_ROOT)}")
