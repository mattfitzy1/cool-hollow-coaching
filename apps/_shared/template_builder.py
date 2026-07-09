"""
Shared styling helper for Business Without You client-input Excel templates.

Every milestone tool ships a branded template with three kinds of tabs:

- Instructions: how to use it, every column explained in plain English,
  grouped by tab for multi-tab templates.
- Examples: a filled-in reference copy of each data tab, clearly marked,
  never read by the tools.
- Data tab(s): a title block, a styled header row, then clean validated
  blank rows. No example rows mixed into the data area, so what a client
  types is exactly what the tool reads.

Column headers use spaces ("Hours Per Week"), not underscores, because the
analysis.py loaders normalize headers to lowercase-with-underscores before
checking them. The shared reader in client_upload.py finds the header row
below the title block, so the branding never breaks the upload.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Brand palette (context/brand.md, approved June 2026): black ink, white
# paper, gold as the single accent, champagne for subtle highlights only.
INK = "1A1A1A"
ACCENT = "C8A227"
ACCENT2 = "E8C766"
PAPER = "FFFFFF"
COMPANY = "Cool Hollow Coaching"
PROGRAM = "Business Without You"

TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
SUBTITLE_FONT = Font(name="Calibri", size=12, bold=True, color=ACCENT2)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=11, color=INK)
NOTE_FONT = Font(name="Calibri", size=10, italic=True, color="5A5A5A")
WARNING_FONT = Font(name="Calibri", size=10, bold=True, italic=True, color=INK)
EXAMPLE_FONT = Font(name="Calibri", size=11, italic=True, color="6B6B6B")
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color=INK)

TITLE_FILL = PatternFill("solid", fgColor=INK)
HEADER_FILL = PatternFill("solid", fgColor=INK)
ACCENT_FILL = PatternFill("solid", fgColor=ACCENT2)
EXAMPLE_FILL = PatternFill("solid", fgColor="EFEFEF")

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def add_note_with_warning(ws, row: int, span: int, lead_text: str, warning_text: str,
                          lead_height: int = 30, warning_height: int = 16, start_col: int = 1):
    """Writes a plain italic note on `row`, then a bold warning line directly
    below it on its own row. Two cells with a single Font each, not rich
    text: openpyxl's per-run rich-text formatting does not reliably survive
    export from a merged, wrap_text cell in Excel/LibreOffice, so a single
    bold sentence in its own cell is the reliable way to make a warning like
    "the tool never reads this tab" impossible to miss.
    Returns the next free row after the warning line."""
    note_cell = ws.cell(row=row, column=start_col, value=lead_text)
    note_cell.font = NOTE_FONT
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col + span - 1)
    ws.row_dimensions[row].height = lead_height

    warn_row = row + 1
    warn_cell = ws.cell(row=warn_row, column=start_col, value=warning_text)
    warn_cell.font = WARNING_FONT
    warn_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=warn_row, start_column=start_col, end_row=warn_row, end_column=start_col + span - 1)
    ws.row_dimensions[warn_row].height = warning_height
    return warn_row + 1


def add_title_block(ws, milestone_label: str, tool_name: str, span: int):
    """Two merged title rows: company/program, then milestone and tool name."""
    span = max(span, 2)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    title_cell = ws.cell(row=1, column=1, value=f"{COMPANY} · {PROGRAM}")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(2, span + 1):
        ws.cell(row=1, column=col).fill = TITLE_FILL
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    subtitle_cell = ws.cell(row=2, column=1, value=f"{milestone_label}: {tool_name}")
    subtitle_cell.font = SUBTITLE_FONT
    subtitle_cell.fill = TITLE_FILL
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(2, span + 1):
        ws.cell(row=2, column=col).fill = TITLE_FILL
    ws.row_dimensions[2].height = 22

    ws.row_dimensions[3].height = 6


def add_header_row(ws, headers: list, row: int = 4, start_col: int = 1):
    for i, header in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 32


def style_data_rows(ws, start_row: int, end_row: int, start_col: int, end_col: int):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.fill = PatternFill("solid", fgColor=PAPER)


def apply_number_formats(ws, formats: list, start_row: int, end_row: int, start_col: int = 1):
    """formats: one Excel number-format string per column, None to skip."""
    for i, fmt in enumerate(formats):
        if not fmt:
            continue
        for r in range(start_row, end_row + 1):
            ws.cell(row=r, column=start_col + i).number_format = fmt


def set_column_widths(ws, widths: list, start_col: int = 1):
    for i, width in enumerate(widths):
        ws.column_dimensions[get_column_letter(start_col + i)].width = width


def set_tab_color(ws, color: str = ACCENT):
    ws.sheet_properties.tabColor = color


def add_list_validation(ws, options: list, cell_range: str, hint: str = ""):
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    dv.error = f"Choose one of: {', '.join(options)}"
    dv.errorTitle = "Invalid entry"
    if hint:
        dv.prompt = hint
        dv.promptTitle = "Pick from the list"
        dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(cell_range)
    return dv


def add_range_validation(ws, low: int, high: int, cell_range: str, hint: str = ""):
    dv = DataValidation(type="whole", operator="between", formula1=low, formula2=high, allow_blank=True)
    dv.error = f"Enter a whole number from {low} to {high}."
    dv.errorTitle = "Out of range"
    if hint:
        dv.prompt = hint
        dv.promptTitle = f"Rate {low} to {high}"
        dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(cell_range)
    return dv


def build_instructions_sheet(wb: Workbook, milestone_label: str, tool_name: str,
                             intro: str, sections: list):
    """sections: list of (tab_name, [(column_name, explanation), ...]) tuples.
    Single-tab templates pass one section; its tab name still shows so the
    client always knows exactly which tab a column lives on."""
    ws = wb.create_sheet("Instructions", 0)
    add_title_block(ws, milestone_label, tool_name, span=2)
    set_tab_color(ws, INK)

    ws.cell(row=5, column=1, value="How to use this template").font = SECTION_FONT
    intro_cell = ws.cell(row=6, column=1, value=intro)
    intro_cell.font = BODY_FONT
    intro_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=2)
    ws.row_dimensions[6].height = 68

    row = add_note_with_warning(
        ws, 7, span=2,
        lead_text="Three steps: fill in the data tab(s), save this file, then upload this "
                  "same file into the tool. The Examples and Starter Ideas tabs are for reference.",
        warning_text="The tool never reads them, only type your real data on the data tab(s).",
        lead_height=32, warning_height=16,
    )
    row += 1
    for tab_name, column_notes in sections:
        section_cell = ws.cell(row=row, column=1, value=f'On the "{tab_name}" tab')
        section_cell.font = Font(name="Calibri", size=11, bold=True, color=INK)
        section_cell.fill = ACCENT_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.row_dimensions[row].height = 20
        row += 1

        ws.cell(row=row, column=1, value="Column").font = HEADER_FONT
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.cell(row=row, column=2, value="What to put here").font = HEADER_FONT
        ws.cell(row=row, column=2).fill = HEADER_FILL
        row += 1

        for col_name, note in column_notes:
            ws.cell(row=row, column=1, value=col_name).font = Font(bold=True, color=INK)
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
            note_cell = ws.cell(row=row, column=2, value=note)
            note_cell.alignment = Alignment(wrap_text=True, vertical="top")
            note_cell.font = BODY_FONT
            ws.row_dimensions[row].height = 32
            row += 1
        row += 1

    set_column_widths(ws, [32, 70])
    return ws


def build_examples_sheet(wb: Workbook, milestone_label: str, tool_name: str, blocks: list):
    """blocks: list of (tab_name, headers, example_rows) tuples. One stacked,
    clearly-labeled example table per data tab. Reference only; the shared
    reader skips this sheet by name."""
    max_span = max(len(headers) for _, headers, _ in blocks)
    ws = wb.create_sheet("Examples")
    add_title_block(ws, milestone_label, tool_name, span=max_span)
    set_tab_color(ws, "6B6B6B")

    add_note_with_warning(
        ws, 4, span=max_span,
        lead_text="Filled-in examples, for reference only. Type your own data on the data tab(s), not here.",
        warning_text="The tool never reads this tab, anything you type here is ignored.",
        lead_height=20, warning_height=16,
    )

    row = 6
    col_widths = [14] * max_span
    for tab_name, headers, example_rows in blocks:
        label = ws.cell(row=row, column=1, value=f'Example for the "{tab_name}" tab')
        label.font = Font(name="Calibri", size=11, bold=True, color=INK)
        label.fill = ACCENT_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        ws.row_dimensions[row].height = 20
        row += 1

        add_header_row(ws, headers, row=row)
        row += 1

        for example in example_rows:
            for c, val in enumerate(example):
                cell = ws.cell(row=row, column=1 + c, value=val)
                cell.font = EXAMPLE_FONT
                cell.fill = EXAMPLE_FILL
                cell.border = THIN_BORDER
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
        row += 2

        for i, header in enumerate(headers):
            col_widths[i] = max(col_widths[i], min(len(str(header)) + 6, 38))

    set_column_widths(ws, col_widths)
    return ws


def build_starter_sheet(wb: Workbook, milestone_label: str, tool_name: str,
                        note: str, groups: list):
    """groups: list of (group_label, [item, ...]) tuples. A scannable menu of
    the most common items owners face, to copy into the data tab only when
    true for their business. Names only, never pre-filled ratings or hours,
    the honesty has to come from the owner. Reference only; the shared
    reader skips this sheet by name."""
    ws = wb.create_sheet("Starter Ideas")
    add_title_block(ws, milestone_label, tool_name, span=2)
    set_tab_color(ws, ACCENT2)

    add_note_with_warning(
        ws, 4, span=2,
        lead_text=note.rstrip(),
        warning_text="The tool never reads this tab, copy items onto the data tab to use them.",
        lead_height=44, warning_height=16,
    )

    row = 6
    for group_label, items in groups:
        label = ws.cell(row=row, column=1, value=group_label)
        label.font = Font(name="Calibri", size=11, bold=True, color=INK)
        label.fill = ACCENT_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.row_dimensions[row].height = 20
        row += 1
        for item in items:
            cell = ws.cell(row=row, column=1, value=item)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            ws.cell(row=row, column=2).border = THIN_BORDER
            row += 1
        row += 1

    set_column_widths(ws, [62, 24])
    return ws


def new_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    return wb
